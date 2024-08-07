# openpyxl zum Arbeiten mit Excel Dateien
import openpyxl

def remove_illegal_characters(excel_data):
    """
    Entfernt Zeichen, die von Openpyxl nicht unterstützt werden und somit nicht in der Excel-Datei verwendet werden
    können aus den Informationen, welche in die Excel-Datei übernommen werden sollen
    https://www.w3schools.com/python/ref_func_ord.asp

    Args:
        excel_data (str): Der String, aus dem für openpyxl illegale Zeichen entfernt werden sollen.

    Returns:
        str: Der bereinigte String ohne für openpyxl illegale Zeichen.
    """
    
    # Entferne alle ASCII-Steuerzeichen welche von Openpyxl nicht unterstützt werden mithilfe von ord()
    return ''.join(char for char in excel_data if ord(char) in range(32, 127))

def find_first_empty_row(sheet, columns):
    """
    Findet die erste leere Zeile in angegebenen Spalten eines Excel-Arbeitsblatts zur Festlegung eines Startpunktes
    zum Einfügen der relevanten Informationen.
    Dies ist nötig, um so flexibel wie möglich mit jeder möglichen Excel-Tabelle arbeiten zu können
    https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html#openpyxl.worksheet.worksheet.Worksheet

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): Das Arbeitsblatt, in dem gesucht wird.
        columns (list): Eine Liste von Spaltennummern, in denen nach der ersten leeren Zeile gesucht wird.

    Returns:
        int: Die Nummer der ersten leeren Zeile, die gefunden wurde.
    """
    for row in range(1, sheet.max_row + 1):
        if all(sheet.cell(row=row, column=col).value is None for col in columns):
            return row
    return sheet.max_row + 1

def update_excel_with_extracted_data(excel_path, extracted_data):
    """
    Aktualisiert die Excel-Datei mit den vorher durch das Modul "pdf_processing" extrahierten Daten aus den PDFs.
    https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html#openpyxl.worksheet.worksheet.Worksheet.cell#
    https://openpyxl.readthedocs.io/en/stable/tutorial.html

    Args:
        excel_path (str): Pfad zur Excel-Datei.
        extracted_data (list): Liste von Tupeln mit den extrahierten Informationen.

    Returns:
        workbook: Das aktualisierte Workbook-Objekt.
    """

    # Öffnen dem angegebenen Arbeitsblatt (sheet) der angegebenen Excel-Datei mithilfe von openpyxl
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook['relevantInfo']

    # Finde die erste leere Zeile in den Spalten A (Paper), C (location coordinates) und D (Area name)
    start_row = find_first_empty_row(worksheet, [2, 6, 10])

    # Trage die aus den PDFs extrahierten Informationen in die Excel-Datei ein
    for i, (pdf_basename, coordinates, lines_with_coordinates, drought_quantified, drought_quantification_keywords, study_type, forest_type, analyzed_years, drought_years) in enumerate(extracted_data):
        worksheet.cell(row=start_row + i, column=1, value=pdf_basename)

        # Kopiere, falls vorhanden, die Koordinaten immer in Spalte C (location coordinates)
        if coordinates != 'Keine Koordinaten gefunden' and len(coordinates.split(', ')) > 1:
            unique_coordinates = ', '.join(sorted(set(coordinates.split(', '))))
            worksheet.cell(row=start_row + i, column=3, value=unique_coordinates)

            # Kopiere, falls vorhanden, die Kontextzeilen der gefundenen Koordinaten immer in Spalte D (Area name)
            worksheet.cell(row=start_row + i, column=4, value=remove_illegal_characters(lines_with_coordinates))

        # Kopiere, falls keine Koordinaten gefunden wurden die Information darüber immer in Spalte C (location coordinates)
        else:
            worksheet.cell(row=start_row + i, column=3, value=coordinates)
            # Kopiere, falls keine Koordinaten gefunden, die Kontextzeilen den gefundenen Bereich der Studie immer in Spalte D (Area name)
            worksheet.cell(row=start_row + i, column=4, value=remove_illegal_characters(lines_with_coordinates))

        # Kopiere, falls ein Schlüsselwort zur Definition von Dürre gefunden wurden die Information darüber wie, immer in Spalte J (how was drought quantified)
        if drought_quantified:
            worksheet.cell(row=start_row + i, column=10, value=remove_illegal_characters(drought_quantified))

        # Kopiere, falls ein Schlüsselwort zum Studientyp gefunden wurde, kopiere den Wert immer in Spalte H (study type)
        if study_type:
            # Konvertiere die Liste der Studientypen in eine Zeichenkette
            study_type_str = ', '.join(study_type)
            worksheet.cell(row=start_row + i, column=8, value=study_type_str)

        # Kopiere, falls ein Studientyp gefunden wurde den Wert immer in Spalte G (ecosystem type)
        if forest_type:
            # Konvertiere die Liste der Studientypen in eine Zeichenkette
            forest_type_str = ', '.join(forest_type)
            worksheet.cell(row=start_row + i, column=7, value=forest_type_str)

        # Kopiere, falls ein Studientyp gefunden wurde den Wert immer in Spalte E (ecosystem type)
        if analyzed_years:
            # Konvertiere die Liste der Studientypen in eine Zeichenkette
            analyzed_years_str = ', '.join(analyzed_years)
            worksheet.cell(row=start_row + i, column=5, value=analyzed_years_str)

        # Kopiere, falls Jahre mit Dürre gefunden wurde den Wert immer in Spalte F (time period with drought (if mentioned))
        if drought_years:
            # Konvertiere die Liste der Studientypen in eine Zeichenkette
            drought_years_str = ', '.join(drought_years)
            worksheet.cell(row=start_row + i, column=6, value=drought_years_str)


    # Speichere die durchgeführten Änderungen in der Excel-Datei
    workbook.save(excel_path)
